from openpyxl import Workbook
wb = Workbook()
# wb.active는 활성화된 시트 정보를 가져옴
ws = wb.create_sheet() # 새로운 Sheet 생성
ws.title = "MySheet" # 시트 이름 변경
ws.sheet_properties.tabColor="ff66ff" # RGB 형태의 값 삽입. #은 제외

# Sheet, Mysheet, YourSheet
ws1 = wb.create_sheet("YourSheet")# 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 시트 이름 출력

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws) # wb 내에 있는 ws 복사
target.title = "Copied Sheet"

wb.save("sample.xlsx")