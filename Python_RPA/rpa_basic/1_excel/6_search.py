# from openpyxl import load_workbook
# wb = load_workbook("sample.xlsx")
# ws = wb.active

# for row in ws.iter_rows(min_row=2):
#     # 번호, 영어, 수학
#     if int(row[1].value) > 80:
#         print(row[0].value, "번 학생은 영어 천재")

# for row in ws.iter_rows(max_row=1): # 데이터를 비교 후 찾아서 값을 바꿔줌
#     for cell in row:
#         if cell.value == "영어":
#             cell.value="컴퓨터"


# wb.save("sample_modified.xlsx")

''' 10.삽입부터 시작. 1:07:19'''
# 지정한 위치에 한줄 삽입
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.insert_rows(8) # 8번째 줄이 비워짐
# ws.insert_rows(8,5) # 8번째 줄 위치로부터 5줄 빈줄 추가
# ws.insert_cols(2) # B번째 열이 비워짐
ws.insert_cols(2,3) # B번째 열로부터 3열 추가
wb.save("sample_insert_cols.xlsx")