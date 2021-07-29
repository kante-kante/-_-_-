from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
# 번호 (국어) 영어 수학

# ws.move_range("B1:C11", rows=0, cols=1) # 영어 수학 점수를 그대로 한 열 옆으로 이동
# ws["B1"].value="국어" # B1셀에 '국어' 입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1) # 수학 값을 6번째 줄 왼쪽 이동

wb.save("sample_korean.xlsx")